import tkinter as tk
import tkinter.messagebox
import threading
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import queue
import openpyxl
import os


class Main_logic():

    def __init__(self, queue, gui):

        super().__init__()
        self.driver = webdriver
        self.driver = webdriver.Chrome(executable_path="chromedriver.exe")
        self.num_apt = 0
        self.queue = queue
        self.gui = gui

# aptname을 juso.go.kr에서 검색 하도록 하는 함수
    def juso_input(self, aptname):
        self.aptname = aptname
        self.driver.get('http://www.juso.go.kr/openIndexPage.do')
        if aptname is not None:
            self.driver.find_element_by_id(
                'inputSearchAddr').send_keys(self.aptname + Keys.RETURN)
            self.num_apt = self.driver.find_elements_by_class_name('count')[
                0].text

# 동, 층, 호수 검색을 위해 사용자가 클릭 후, option이 display되는 상태 확인을 위한 함수(Thread)
    def detecting_choice(self):
        old_elements = []

        while True:

            if self.queue.qsize() == 0:

                element = WebDriverWait(self.driver, 6000).until(
                    EC.presence_of_element_located((
                        By.XPATH, "//*[contains(@id,'popAddrDetail2_') and @style='display: block;']"
                    )))
                if element is not None:
                    elements = self.driver.find_elements_by_xpath(
                        "//*[contains(@id,'popAddrDetail2_') and @style='display: block;']"
                    )
                    if len(old_elements) < len(elements):

                        old_elements = elements
                        final_element = elements[-1]
                        selected_full_id = final_element.get_attribute('id')
                        selected_slice_id = selected_full_id[15:]
                        self.queue.put(selected_slice_id)
                        self.gui.popup_excel()
                    else:
                        old_elements = elements
            else:
                print("sleep")
                time.sleep(0.1)

# APT의 동, 호수 작성한 Excel 파일을 바탕화면에 추출
    def excel_extract(self, queue):

        wb = openpyxl.Workbook()
        sheet1 = wb.active
        sheet1['A1'] = '동'
      #  sheet1['B1'] = '층'
        sheet1['B1'] = '호'
        sheet_row = 1

        selected_slice_id = str(queue.get())
        selected_dong_id = "sel_dong"+selected_slice_id

        dong_list = str(self.driver.find_element_by_id(
            selected_dong_id).text).split('\n')
        dong_list.pop(0)
        len_dong = len(dong_list)

        for number_floor in range(2, len_dong+2):

            dong_xpath = '//*[@id="sel_dong{queue_input}"]/option[{num_input}]'.format(
                queue_input=selected_slice_id, num_input=number_floor)
            final_dong = self.driver.find_element_by_xpath(
                dong_xpath).text.replace('동', '')
            self.driver.find_element_by_xpath(dong_xpath).click()
            time.sleep(0.5)
            floor_list = str(self.driver.find_element_by_id(
                "sel_dong_floor"+selected_slice_id).text).split('\n')
            floor_list.pop(0)
            len_floor = len(floor_list)
            print(len_floor)

            for number_ho in range(2, len_floor+2):
                # try:
                floor_xpath = '//*[@id="sel_dong_floor{queue_input}"]/option[{num_input}]'.format(
                    queue_input=selected_slice_id, num_input=number_ho)
                self.driver.find_element_by_xpath(floor_xpath).click()
                #final_floor = self.driver.find_element_by_xpath(floor_xpath).text.replace("층","")
                time.sleep(0.5)
                ho_list = str(self.driver.find_element_by_id(
                    "sel_dong_ho"+selected_slice_id).text).split('\n')
                ho_list.pop(0)
                len_ho = len(ho_list)
                # except:
                #    print("오류 발생 동 : " +str(final_dong)+ "층 : "+ str(final_floor) + "분석에서 오류가 났습니다.")
                #    pass

                #    try:
                for final_ho in range(0, len_ho):
                    try:
                        sheet_row = 1 + sheet_row
                        final_ho_name = ho_list[final_ho].replace("호", "")
                        sheet1.cell(row=sheet_row, column=1, value=final_dong)
                        #sheet1.cell(row = sheet_row, column = 2, value = final_floor)
                        sheet1.cell(row=sheet_row, column=2,
                                    value=final_ho_name)
                    except:
                        print("오류 발생 동 : " + str(final_dong)
                              + "호 : " + str(final_ho_name) + "분석에서 오류가 났습니다.")
                        pass

        excel_addr = os.path.join(os.path.join(
            os.environ['USERPROFILE']), 'Desktop') + "/{} 동호수리스트.xlsx".format(self.aptname)
        wb.save(excel_addr)
        wb.close
        print("end")


class Call_ui(tk.Tk):

    def __init__(self, queue):
        self.input_aptname = ""
        self.queue = queue
        self.ml = Main_logic(self.queue, self)
        t1 = threading.Thread(target=self.ml.detecting_choice)
        t1.start()

        tk.Tk.__init__(self)

        self.title("단지 검색")
        self.geometry("400x100")
        self.resizable(False, False)

        self.aptname_input_entry = tk.Entry(self,
                                            width=30
                                            )
        self.aptname_input_entry.place(x=50, y=35)

        self.find_button = tk.Button(self,
                                     text="검색",
                                     width=10,
                                     padx=1,
                                     pady=1,
                                     command=self.click_button
                                     )
        self.find_button.place(x=300, y=45)

        self.mainloop()

    def click_button(self):

        self.input_aptname = self.aptname_input_entry.get()
        self.ml.juso_input(self.input_aptname)

    def popup_excel(self):
        self.msgbox = tk.messagebox.askquestion(
            '요청', '동호수 Excel 추출하시겠습니까?')
        if self.msgbox == 'yes':
            self.ml.excel_extract(queue)

        else:
            self.queue.get()


queue = queue.Queue(maxsize=1)
cu = Call_ui(queue)
